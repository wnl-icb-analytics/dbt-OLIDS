"""
Analyse monthly practice count comparisons: PMCT vs OLIDS.
Generates a detailed markdown report focusing on overall counts and age distribution by life stage.
PMCT is the source of truth for monthly practice counts.
"""

import os
from pathlib import Path
from datetime import datetime
import numpy as np
from dotenv import load_dotenv
from snowflake.snowpark import Session
import pandas as pd

# Load environment variables
load_dotenv()

# Configuration
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

PMCT_QUERY_FILE = SCRIPT_DIR / "pmct_monthly_counts_native_query.sql"
OLIDS_QUERY_FILE = SCRIPT_DIR / "olids_monthly_counts_native_query.sql"
WAREHOUSE = os.getenv('SNOWFLAKE_WAREHOUSE')
ACCOUNT = os.getenv('SNOWFLAKE_ACCOUNT')
USER = os.getenv('SNOWFLAKE_USER')
ROLE = os.getenv('SNOWFLAKE_ROLE')


def parse_sql_file(file_path):
    """Parse SQL file and extract USE statements and the main query."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    lines = content.split('\n')
    use_statements = []
    query_lines = []
    in_query = False
    
    for line in lines:
        line_stripped = line.strip()
        line_upper = line_stripped.upper()
        
        if line_upper.startswith('USE '):
            use_statements.append(line_stripped.rstrip(';'))
            continue
        
        if line_upper.startswith('WITH') or line_upper.startswith('SELECT'):
            in_query = True
        
        if in_query:
            query_lines.append(line)
    
    query = '\n'.join(query_lines).strip().rstrip(';')
    
    if not query or query.upper().startswith('USE'):
        statements = [s.strip() for s in content.split(';')]
        for stmt in statements:
            stmt_upper = stmt.upper()
            if stmt_upper.startswith('WITH') or stmt_upper.startswith('SELECT'):
                query = stmt.strip()
                break
    
    return use_statements, query


def fix_encoding(text):
    """Fix common encoding issues in text."""
    if pd.isna(text):
        return ''
    text_str = str(text)
    # Fix common encoding issues
    replacements = {
        '�': "'",  # Replace replacement character with apostrophe
        '\x92': "'",  # Windows-1252 apostrophe
        '\x93': '"',  # Windows-1252 opening quote
        '\x94': '"',  # Windows-1252 closing quote
        '\u2019': "'",  # Right single quotation mark
        '\u2018': "'",  # Left single quotation mark
        '\u201C': '"',  # Left double quotation mark
        '\u201D': '"',  # Right double quotation mark
    }
    for old, new in replacements.items():
        text_str = text_str.replace(old, new)
    # Fix specific known issues
    text_str = text_str.replace("Wakeman�s", "Wakeman's")
    text_str = text_str.replace("St George�s", "St George's")
    return text_str


def execute_query(session, use_statements, query, query_name, session_role=None):
    """Execute USE statements and SQL query, return results as pandas DataFrame."""
    print(f"Executing {query_name}...")
    try:
        # Execute USE statements individually to ensure role/warehouse/database changes persist
        # Skip USE ROLE if session is already created with the correct role
        for i, use_stmt in enumerate(use_statements):
            use_stmt_upper = use_stmt.upper()
            # Skip USE ROLE if session was created with that role already
            if use_stmt_upper.startswith('USE ROLE') and session_role:
                # Extract role from USE ROLE statement
                role_in_stmt = None
                if '"' in use_stmt:
                    role_in_stmt = use_stmt.split('"')[1]
                elif "'" in use_stmt:
                    role_in_stmt = use_stmt.split("'")[1]
                
                if role_in_stmt and role_in_stmt.upper() == session_role.upper():
                    print(f"  Skipping USE ROLE (session already using {session_role})...")
                    continue
            
            print(f"  Executing USE statement {i+1}/{len(use_statements)}: {use_stmt[:50]}...")
            try:
                session.sql(use_stmt + ';').collect()
                print(f"    ✓ Success")
            except Exception as e:
                # If USE ROLE fails but session was created with correct role, continue
                if use_stmt_upper.startswith('USE ROLE') and session_role:
                    print(f"    Warning: USE ROLE failed but session already has correct role, continuing...")
                    continue
                # If USE WAREHOUSE fails, try to continue (warehouse might not be accessible to this role)
                elif use_stmt_upper.startswith('USE WAREHOUSE'):
                    print(f"    Warning: USE WAREHOUSE failed - {e}")
                    print(f"    Continuing without explicit warehouse (may use default)...")
                    continue
                else:
                    print(f"    ✗ Failed: {e}")
                    print(f"    Statement: {use_stmt}")
                    raise
        
        # Execute the main query (role/warehouse/database from USE statements should persist)
        print(f"  Executing main query...")
        df = session.sql(query).to_pandas()
        # Fix encoding issues in text columns
        text_columns = df.select_dtypes(include=['object']).columns
        for col in text_columns:
            df[col] = df[col].apply(fix_encoding)
        print(f"✓ {query_name}: {len(df)} rows")
        return df
    except Exception as e:
        print(f"✗ Error executing {query_name}: {e}")
        print(f"  USE statements: {use_statements}")
        raise


def prepare_overall_comparison(pmct_df, olids_df):
    """Prepare overall count comparison (ignoring age/life stage)."""
    # Aggregate PMCT data to practice level (sum across life stages)
    pmct_overall = pmct_df.groupby(['PRACTICE_CODE', 'PRACTICE_NAME']).agg({
        'PMCT_TOTAL_COUNT': 'first'  # Should be same for all rows per practice
    }).reset_index()
    
    # Aggregate OLIDS data to practice level (sum across life stages)
    olids_overall = olids_df.groupby(['PRACTICE_CODE', 'PRACTICE_NAME']).agg({
        'OLIDS_TOTAL_COUNT': 'first'  # Should be same for all rows per practice
    }).reset_index()
    
    # Merge
    comparison = pd.merge(
        pmct_overall,
        olids_overall,
        on='PRACTICE_CODE',
        how='outer',
        suffixes=('_PMCT', '_OLIDS')
    )
    
    # Use the first non-null practice name
    comparison['PRACTICE_NAME'] = comparison['PRACTICE_NAME_PMCT'].fillna(
        comparison['PRACTICE_NAME_OLIDS']
    )
    comparison = comparison.drop(columns=['PRACTICE_NAME_PMCT', 'PRACTICE_NAME_OLIDS'])
    
    # Calculate differences
    comparison['DIFFERENCE'] = comparison['OLIDS_TOTAL_COUNT'] - comparison['PMCT_TOTAL_COUNT']
    comparison['DIFFERENCE_PCT'] = (
        comparison['DIFFERENCE'] / comparison['PMCT_TOTAL_COUNT'].replace(0, pd.NA) * 100
    ).fillna(0)
    comparison['ABS_DIFFERENCE_PCT'] = comparison['DIFFERENCE_PCT'].abs()
    
    # Categorise matches
    comparison['MATCH_CATEGORY'] = comparison.apply(
        lambda row: (
            'OLIDS Only' if pd.isna(row['PMCT_TOTAL_COUNT']) else
            'PMCT Only' if pd.isna(row['OLIDS_TOTAL_COUNT']) else
            'Major Difference (±20%+)' if row['ABS_DIFFERENCE_PCT'] >= 20 else
            'Moderate Difference (5-20%)' if row['ABS_DIFFERENCE_PCT'] >= 5 else
            'Good Match (1-5%)' if row['ABS_DIFFERENCE_PCT'] >= 1 else
            'Excellent Match (<1%)'
        ),
        axis=1
    )
    
    return comparison


def prepare_life_stage_comparison(pmct_df, olids_df):
    """Prepare life stage comparison."""
    # Prepare PMCT life stage data
    pmct_life_stage = pmct_df[
        pmct_df['LIFE_STAGE'].notna()
    ].groupby(['PRACTICE_CODE', 'PRACTICE_NAME', 'LIFE_STAGE']).agg({
        'PMCT_COUNT_BY_LIFE_STAGE': 'sum'
    }).reset_index()
    
    # Prepare OLIDS life stage data
    olids_life_stage = olids_df[
        olids_df['LIFE_STAGE'].notna()
    ].groupby(['PRACTICE_CODE', 'PRACTICE_NAME', 'LIFE_STAGE']).agg({
        'OLIDS_COUNT_BY_LIFE_STAGE': 'sum'
    }).reset_index()
    
    # Merge
    comparison = pd.merge(
        pmct_life_stage,
        olids_life_stage,
        on=['PRACTICE_CODE', 'LIFE_STAGE'],
        how='outer',
        suffixes=('_PMCT', '_OLIDS')
    )
    
    # Use the first non-null practice name
    comparison['PRACTICE_NAME'] = comparison['PRACTICE_NAME_PMCT'].fillna(
        comparison['PRACTICE_NAME_OLIDS']
    )
    comparison = comparison.drop(columns=['PRACTICE_NAME_PMCT', 'PRACTICE_NAME_OLIDS'])
    
    # Calculate differences
    comparison['DIFFERENCE'] = (
        comparison['OLIDS_COUNT_BY_LIFE_STAGE'].fillna(0) - 
        comparison['PMCT_COUNT_BY_LIFE_STAGE'].fillna(0)
    )
    comparison['DIFFERENCE_PCT'] = (
        comparison['DIFFERENCE'] / comparison['PMCT_COUNT_BY_LIFE_STAGE'].replace(0, pd.NA) * 100
    ).fillna(0)
    comparison['ABS_DIFFERENCE_PCT'] = comparison['DIFFERENCE_PCT'].abs()
    
    return comparison


def analyse_age_distribution_for_good_matches(overall_df, life_stage_df):
    """Analyse age distribution for practices with good overall matches."""
    # Filter to practices with good/excellent overall matches
    good_matches = overall_df[
        overall_df['MATCH_CATEGORY'].isin(['Excellent Match (<1%)', 'Good Match (1-5%)'])
    ].copy()
    
    good_match_codes = set(good_matches['PRACTICE_CODE'].unique())
    
    # Get life stage data for these practices
    life_stage_good_matches = life_stage_df[
        life_stage_df['PRACTICE_CODE'].isin(good_match_codes)
    ].copy()
    
    # Calculate age distribution percentages for each practice
    results = []
    
    for practice_code in good_match_codes:
        practice_data = life_stage_good_matches[
            life_stage_good_matches['PRACTICE_CODE'] == practice_code
        ].copy()
        
        if len(practice_data) == 0:
            continue
        
        practice_name = practice_data.iloc[0]['PRACTICE_NAME']
        overall_match = good_matches[
            good_matches['PRACTICE_CODE'] == practice_code
        ].iloc[0]['MATCH_CATEGORY']
        
        # Calculate totals for percentage calculations
        pmct_total = practice_data['PMCT_COUNT_BY_LIFE_STAGE'].fillna(0).sum()
        olids_total = practice_data['OLIDS_COUNT_BY_LIFE_STAGE'].fillna(0).sum()
        
        # Calculate percentage distributions
        practice_data['PMCT_PCT'] = (
            practice_data['PMCT_COUNT_BY_LIFE_STAGE'].fillna(0) / pmct_total * 100
            if pmct_total > 0 else 0
        )
        practice_data['OLIDS_PCT'] = (
            practice_data['OLIDS_COUNT_BY_LIFE_STAGE'].fillna(0) / olids_total * 100
            if olids_total > 0 else 0
        )
        
        # Calculate difference in percentage points
        practice_data['PCT_POINT_DIFF'] = practice_data['OLIDS_PCT'] - practice_data['PMCT_PCT']
        practice_data['ABS_PCT_POINT_DIFF'] = practice_data['PCT_POINT_DIFF'].abs()
        
        # Determine if age distribution matches well
        max_pct_diff = practice_data['ABS_PCT_POINT_DIFF'].max()
        if max_pct_diff < 1:
            age_dist_match = 'Excellent (<1pp)'
        elif max_pct_diff < 5:
            age_dist_match = 'Good (1-5pp)'
        elif max_pct_diff < 10:
            age_dist_match = 'Moderate (5-10pp)'
        else:
            age_dist_match = 'Poor (>10pp)'
        
        results.append({
            'PRACTICE_CODE': practice_code,
            'PRACTICE_NAME': practice_name,
            'OVERALL_MATCH': overall_match,
            'AGE_DIST_MATCH': age_dist_match,
            'MAX_PCT_POINT_DIFF': max_pct_diff,
            'LIFE_STAGE_DATA': practice_data
        })
    
    return results


def analyse_overall_comparison(overall_df):
    """Analyse overall comparison statistics."""
    valid_comparison = overall_df[
        (overall_df['PMCT_TOTAL_COUNT'].notna()) & 
        (overall_df['OLIDS_TOTAL_COUNT'].notna())
    ].copy()
    
    total_pmct = valid_comparison['PMCT_TOTAL_COUNT'].sum()
    total_olids = valid_comparison['OLIDS_TOTAL_COUNT'].sum()
    overall_diff = total_olids - total_pmct
    overall_diff_pct = (overall_diff / total_pmct * 100) if total_pmct > 0 else 0
    
    analysis = {
        'total_practices': len(overall_df),
        'excellent_match': len(overall_df[overall_df['MATCH_CATEGORY'] == 'Excellent Match (<1%)']),
        'good_match': len(overall_df[overall_df['MATCH_CATEGORY'] == 'Good Match (1-5%)']),
        'moderate_diff': len(overall_df[overall_df['MATCH_CATEGORY'] == 'Moderate Difference (5-20%)']),
        'major_diff': len(overall_df[overall_df['MATCH_CATEGORY'] == 'Major Difference (±20%+)']),
        'pmct_only': len(overall_df[overall_df['MATCH_CATEGORY'] == 'PMCT Only']),
        'olids_only': len(overall_df[overall_df['MATCH_CATEGORY'] == 'OLIDS Only']),
        'total_pmct': total_pmct,
        'total_olids': total_olids,
        'overall_diff': overall_diff,
        'overall_diff_pct': overall_diff_pct,
        'df': overall_df.copy()
    }
    
    return analysis


def format_pct_diff(value, decimals=1):
    """Format percentage difference, handling NaN and infinity."""
    if pd.isna(value) or np.isinf(value):
        return "—"
    return f"{value:+.{decimals}f}%"


def export_to_excel(age_dist_analysis, output_file):
    """Export the age distribution summary table to Excel."""
    try:
        import openpyxl
    except ImportError:
        print("⚠ Warning: openpyxl not installed. Skipping Excel export.")
        print("  Install with: pip install openpyxl")
        return False
    
    if len(age_dist_analysis) == 0:
        print("⚠ Warning: No age distribution data to export.")
        return False
    
    try:
        # Sort by age distribution match quality, then by max percentage point difference
        sorted_results = sorted(age_dist_analysis, key=lambda x: (
            {'Excellent (<1pp)': 1, 'Good (1-5pp)': 2, 'Moderate (5-10pp)': 3, 'Poor (>10pp)': 4}[x['AGE_DIST_MATCH']],
            x['MAX_PCT_POINT_DIFF']
        ))
        
        # Get all unique life stages to create columns (sorted by age range)
        all_life_stages = set()
        for result in age_dist_analysis:
            for _, row in result['LIFE_STAGE_DATA'].iterrows():
                all_life_stages.add(row['LIFE_STAGE'])
        
        # Sort life stages by extracting the numeric start of the age range
        def get_age_start(life_stage):
            if ' (' in life_stage:
                age_part = life_stage.split(' (')[0]
                if '-' in age_part:
                    return int(age_part.split('-')[0])
                elif '+' in age_part:
                    return int(age_part.replace('+', ''))
            return 999  # Put unknown at end
        
        all_life_stages = sorted(all_life_stages, key=get_age_start)
        
        # Build DataFrame
        rows = []
        for result in sorted_results:
            # Determine indicator
            if result['AGE_DIST_MATCH'] == 'Excellent (<1pp)':
                indicator = '⭐'
            elif result['AGE_DIST_MATCH'] == 'Good (1-5pp)':
                indicator = '✅'
            elif result['AGE_DIST_MATCH'] == 'Moderate (5-10pp)':
                indicator = '⚠️'
            else:
                indicator = '❌'
            
            # Create a dictionary for quick lookup
            life_stage_dict = {}
            for _, row in result['LIFE_STAGE_DATA'].iterrows():
                life_stage_dict[row['LIFE_STAGE']] = {
                    'pmct_pct': row['PMCT_PCT'],
                    'olids_pct': row['OLIDS_PCT']
                }
            
            # Build row dictionary - match markdown format exactly
            row_dict = {
                'Indicator': indicator,
                'Practice Code': result['PRACTICE_CODE'],
                'Practice Name': result['PRACTICE_NAME'],
                'Overall Match': result['OVERALL_MATCH'],
                'Age Dist Match': result['AGE_DIST_MATCH'],
                'Max Diff': f"{result['MAX_PCT_POINT_DIFF']:.1f}pp"
            }
            
            # Add life stage columns - use exact same format as markdown
            for life_stage in all_life_stages:
                if life_stage in life_stage_dict:
                    pmct_pct = life_stage_dict[life_stage]['pmct_pct']
                    olids_pct = life_stage_dict[life_stage]['olids_pct']
                    # Store as numbers for Excel (not strings with %)
                    row_dict[f"{life_stage} PMCT%"] = pmct_pct
                    row_dict[f"{life_stage} OLIDS%"] = olids_pct
                else:
                    row_dict[f"{life_stage} PMCT%"] = None
                    row_dict[f"{life_stage} OLIDS%"] = None
            
            rows.append(row_dict)
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Export to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Age Distribution Summary', index=False)
            
            # Get the worksheet to format
            worksheet = writer.sheets['Age Distribution Summary']
            
            # Format percentage columns and auto-adjust column widths
            from openpyxl.styles import NamedStyle, Font
            from openpyxl.utils import get_column_letter
            
            # Format header row
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
            
            # Format percentage columns and adjust widths
            for idx, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(idx)
                max_length = max(
                    df[col].astype(str).map(len).max() if col in df.columns else 0,
                    len(str(col))
                )
                # Set a reasonable max width
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # Format percentage columns
                if 'PMCT%' in col or 'OLIDS%' in col:
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=idx)
                        if cell.value is not None:
                            cell.number_format = '0.0"%"'
            
            # Freeze first row and first 3 columns
            worksheet.freeze_panes = 'D2'
        
        return True
    except Exception as e:
        print(f"⚠ Error during Excel export: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_markdown_report(overall_analysis, life_stage_df, age_dist_analysis, output_file):
    """Generate a focused markdown report on practices with good matches."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("# PMCT vs OLIDS Monthly Practice Counts Comparison\n\n")
        f.write(f"**Generated:** {timestamp}\n\n")
        f.write("**Source of Truth:** PMCT Monthly Practice Counts (2025-11-01)\n")
        f.write("**Target Date:** 2025-11-01 | **Sub ICB Location Code:** 93C | **All Ages Included**\n\n")
        
        if len(age_dist_analysis) > 0:
            # Summary statistics
            excellent_age_dist = len([r for r in age_dist_analysis if r['AGE_DIST_MATCH'] == 'Excellent (<1pp)'])
            good_age_dist = len([r for r in age_dist_analysis if r['AGE_DIST_MATCH'] == 'Good (1-5pp)'])
            moderate_age_dist = len([r for r in age_dist_analysis if r['AGE_DIST_MATCH'] == 'Moderate (5-10pp)'])
            poor_age_dist = len([r for r in age_dist_analysis if r['AGE_DIST_MATCH'] == 'Poor (>10pp)'])
            
            f.write(f"## {len(age_dist_analysis)} Practices with Good Overall Matches\n\n")
            f.write("This report focuses on practices where overall registration counts match well between PMCT and OLIDS, ")
            f.write("and validates that their age distributions also align. This confirms patients have valid DOB and are correctly categorised.\n\n")
            
            f.write("### Age Distribution Match Quality Summary\n\n")
            f.write("| Age Distribution Match | Count | Percentage |\n")
            f.write("|------------------------|-------|------------|\n")
            f.write(f"| ⭐ Excellent (<1pp) | {excellent_age_dist} | {excellent_age_dist/len(age_dist_analysis)*100:.1f}% |\n")
            f.write(f"| ✅ Good (1-5pp) | {good_age_dist} | {good_age_dist/len(age_dist_analysis)*100:.1f}% |\n")
            f.write(f"| ⚠️ Moderate (5-10pp) | {moderate_age_dist} | {moderate_age_dist/len(age_dist_analysis)*100:.1f}% |\n")
            f.write(f"| ❌ Poor (>10pp) | {poor_age_dist} | {poor_age_dist/len(age_dist_analysis)*100:.1f}% |\n\n")
            
            f.write("**Match Definitions:**\n")
            f.write("- **Overall Match:** Based on total count difference (Excellent <1%, Good 1-5%)\n")
            f.write("- **Age Dist Match:** Maximum percentage point difference across all life stages (Excellent <1pp, Good 1-5pp)\n")
            f.write("- **Age Bands:** 0-4 (Young Children), 5-11 (Children), 12-17 (Teenagers), 18-24 (Young Adults), ")
            f.write("25-44 (Adults), 45-64 (Middle-Aged Adults), 65-74 (Elderly), 75-84 (Very Elderly), 85+ (Very Elderly)\n\n")
            f.write("---\n\n")
            
            # Sort by age distribution match quality, then by max percentage point difference
            sorted_results = sorted(age_dist_analysis, key=lambda x: (
                {'Excellent (<1pp)': 1, 'Good (1-5pp)': 2, 'Moderate (5-10pp)': 3, 'Poor (>10pp)': 4}[x['AGE_DIST_MATCH']],
                x['MAX_PCT_POINT_DIFF']
            ))
            
            # Get all unique life stages to create columns (sorted by age range)
            all_life_stages = set()
            for result in age_dist_analysis:
                for _, row in result['LIFE_STAGE_DATA'].iterrows():
                    all_life_stages.add(row['LIFE_STAGE'])
            
            # Sort life stages by extracting the numeric start of the age range
            def get_age_start(life_stage):
                if ' (' in life_stage:
                    age_part = life_stage.split(' (')[0]
                    if '-' in age_part:
                        return int(age_part.split('-')[0])
                    elif '+' in age_part:
                        return int(age_part.replace('+', ''))
                return 999  # Put unknown at end
            
            all_life_stages = sorted(all_life_stages, key=get_age_start)
            
            # Create header with full life stage names
            header = "| Indicator | Practice Code | Practice Name | Overall Match | Age Dist Match | Max Diff |"
            for life_stage in all_life_stages:
                # Use full life stage name for column header
                if ' (' in life_stage:
                    # Extract full description, e.g., "0-4 (Young Children)" -> "0-4 (Young Children)"
                    header += f" {life_stage} PMCT% | {life_stage} OLIDS% |"
                else:
                    header += f" {life_stage} PMCT% | {life_stage} OLIDS% |"
            header += "\n"
            
            separator = "|-----------|---------------|---------------|---------------|----------------|----------|"
            for _ in all_life_stages:
                separator += "------------|-------------|"
            separator += "\n"
            
            f.write(header)
            f.write(separator)
            
            for result in sorted_results:
                # Determine indicator
                if result['AGE_DIST_MATCH'] == 'Excellent (<1pp)':
                    indicator = '⭐'
                elif result['AGE_DIST_MATCH'] == 'Good (1-5pp)':
                    indicator = '✅'
                elif result['AGE_DIST_MATCH'] == 'Moderate (5-10pp)':
                    indicator = '⚠️'
                else:
                    indicator = '❌'
                
                practice_name = str(result['PRACTICE_NAME']).replace('|', '\\|')
                
                # Create a dictionary for quick lookup
                life_stage_dict = {}
                for _, row in result['LIFE_STAGE_DATA'].iterrows():
                    life_stage_dict[row['LIFE_STAGE']] = {
                        'pmct_pct': row['PMCT_PCT'],
                        'olids_pct': row['OLIDS_PCT']
                    }
                
                # Build row
                row = (f"| {indicator} | {result['PRACTICE_CODE']} | {practice_name} | "
                      f"{result['OVERALL_MATCH']} | {result['AGE_DIST_MATCH']} | "
                      f"{result['MAX_PCT_POINT_DIFF']:.1f}pp |")
                
                for life_stage in all_life_stages:
                    if life_stage in life_stage_dict:
                        pmct_pct = life_stage_dict[life_stage]['pmct_pct']
                        olids_pct = life_stage_dict[life_stage]['olids_pct']
                        row += f" {pmct_pct:.1f}% | {olids_pct:.1f}% |"
                    else:
                        row += " — | — |"
                
                f.write(row + "\n")
            f.write("\n")


def main():
    """Execute analysis and generate report."""
    print("="*80)
    print("PMCT vs OLIDS Monthly Practice Counts Comparison")
    print("="*80)
    
    # Read SQL files
    print("\nReading SQL files...")
    pmct_use, pmct_query = parse_sql_file(PMCT_QUERY_FILE)
    olids_use, olids_query = parse_sql_file(OLIDS_QUERY_FILE)
    print("✓ SQL files read")
    
    # Create separate sessions for PMCT (ENGINEER role) and OLIDS (env role)
    print("\nConnecting to Snowflake...")
    
    # Session for PMCT - use ENGINEER role (extracted from PMCT query USE statements)
    pmct_role = None
    for use_stmt in pmct_use:
        if use_stmt.upper().startswith('USE ROLE'):
            # Extract role name from USE ROLE statement
            pmct_role = use_stmt.split('"')[1] if '"' in use_stmt else use_stmt.split("'")[1] if "'" in use_stmt else None
            break
    
    pmct_session = None
    olids_session = None
    
    try:
        # Create PMCT session with ENGINEER role
        # Don't specify warehouse - ENGINEER role may have a different default warehouse
        if pmct_role:
            print(f"Creating PMCT session with {pmct_role} role...")
            pmct_connection_params = {
                "account": ACCOUNT,
                "user": USER,
                "authenticator": "externalbrowser",
                "role": pmct_role
                # Note: Not specifying warehouse - will use ENGINEER role's default warehouse
            }
            pmct_session = Session.builder.configs(pmct_connection_params).create()
            print("✓ PMCT session connected")
        else:
            # Fallback: use env role and try to switch
            print("Creating PMCT session with env role (will switch to ENGINEER)...")
            pmct_connection_params = {
                "account": ACCOUNT,
                "user": USER,
                "authenticator": "externalbrowser",
                "warehouse": WAREHOUSE,
                "role": ROLE
            }
            pmct_session = Session.builder.configs(pmct_connection_params).create()
            print("✓ PMCT session connected")
        
        # Create OLIDS session with env role
        print("Creating OLIDS session with env role...")
        olids_connection_params = {
            "account": ACCOUNT,
            "user": USER,
            "authenticator": "externalbrowser",
            "warehouse": WAREHOUSE,
            "role": ROLE
        }
        olids_session = Session.builder.configs(olids_connection_params).create()
        print("✓ OLIDS session connected")
        
        # Execute queries with appropriate session roles
        pmct_df = execute_query(pmct_session, pmct_use, pmct_query, "PMCT Monthly Counts", session_role=pmct_role)
        olids_df = execute_query(olids_session, olids_use, olids_query, "OLIDS Monthly Counts", session_role=ROLE)
        
        # Prepare comparisons
        print("\nPreparing comparisons...")
        overall_comparison = prepare_overall_comparison(pmct_df, olids_df)
        life_stage_comparison = prepare_life_stage_comparison(pmct_df, olids_df)
        
        # Analyse results
        print("Analysing results...")
        overall_analysis = analyse_overall_comparison(overall_comparison)
        age_dist_analysis = analyse_age_distribution_for_good_matches(overall_comparison, life_stage_comparison)
        
        # Generate report
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_file = OUTPUT_DIR / f"pmct_olids_comparison_{timestamp}.md"
        generate_markdown_report(overall_analysis, life_stage_comparison, age_dist_analysis, report_file)
        
        print(f"\n✓ Analysis complete")
        print(f"✓ Report saved to: {report_file}")
        
        # Export summary table to Excel
        excel_file = OUTPUT_DIR / f"pmct_olids_comparison_{timestamp}.xlsx"
        if export_to_excel(age_dist_analysis, excel_file):
            if excel_file.exists():
                print(f"✓ Excel file saved to: {excel_file}")
            else:
                print("⚠ Warning: Excel export completed but file not found")
        else:
            print("⚠ Warning: Excel export was skipped or failed")
        
        # Save CSV files for further analysis
        overall_comparison.to_csv(OUTPUT_DIR / "pmct_olids_overall_comparison.csv", index=False)
        life_stage_comparison.to_csv(OUTPUT_DIR / "pmct_olids_life_stage_comparison.csv", index=False)
        print(f"✓ CSV files saved to: {OUTPUT_DIR}")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        raise
    finally:
        if pmct_session:
            pmct_session.close()
            print("PMCT session closed")
        if olids_session:
            olids_session.close()
            print("OLIDS session closed")
        print("\nAll connections closed")


if __name__ == '__main__':
    main()

